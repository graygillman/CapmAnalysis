#Standard Libraries
import os
from datetime import date

#Third Party Libraries
import numpy as np
import pandas as pd
import requests
import statsmodels.api as sm
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import plotly.graph_objects as go


class CAPMAnalysis:
    def __init__(self, ticker, benchmark, risk_free, daily_or_monthly):
        self.ticker = ticker.upper()
        self.benchmark = benchmark.upper()
        self.risk_free = risk_free.upper()
        self.daily_or_monthly = daily_or_monthly.upper()
        self.interval = "1d" if self.daily_or_monthly == 'D' else '1mo'
        self.days_or_months = 365 if self.daily_or_monthly == 'D' else 12
        self.word = 'Daily' if self.daily_or_monthly == 'D' else 'Monthly'
        self.periods_in_years = [1, 2, 3, 4, 5] if self.daily_or_monthly == 'D' else [1, 5, 10, 20]
        self.period = '5y' if self.interval == '1d' else 'max'
        self.monthly_returns = None
        self.results_df = None
        self.today = date.today()
    
    def download_data(self):
        self.ticker_df = yf.download(tickers=self.ticker, period=self.period, interval=self.interval)
        self.benchmark_df = yf.download(tickers=self.benchmark, period=self.period, interval=self.interval)
        self.rf_df = yf.download(tickers=self.risk_free, period=self.period, interval=self.interval)
    
    def calculate_returns(self):
        self.monthly_returns = pd.concat([self.ticker_df['Adj Close'], self.benchmark_df['Adj Close']], axis=1)
        self.monthly_returns.columns = [self.ticker, self.benchmark]
        self.monthly_returns = self.monthly_returns.dropna() #Drops any values if the stock or benchmark doesnt have data that goes more than 5 years back

        self.monthly_returns = self.monthly_returns.pct_change(1).dropna()
        self.monthly_returns[f'{self.risk_free} (RF)'] = (self.rf_df['Adj Close'] / 100) / self.days_or_months
        self.monthly_returns[f'{self.ticker} - RF'] = self.monthly_returns[self.ticker] - self.monthly_returns[f'{self.risk_free} (RF)']
        self.monthly_returns[f'{self.benchmark} - RF'] = self.monthly_returns[self.benchmark] - self.monthly_returns[f'{self.risk_free} (RF)']

        self.monthly_returns.index = pd.to_datetime(self.monthly_returns.index)
        self.monthly_returns.index = self.monthly_returns.index.date
        self.first_date = self.monthly_returns.index.min()
        self.last_date = self.monthly_returns.index.max()
        data_range_msg = f"Data Range: {self.first_date} to {self.last_date}"
        self.years_difference = (self.last_date - self.first_date).days / 365.25  

        return data_range_msg


    def perform_regression(self):
        X = sm.add_constant(self.monthly_returns[f'{self.benchmark} - RF'])
        y = self.monthly_returns[f'{self.ticker} - RF']
        model = sm.OLS(y, X).fit()
        
        self.results_df = pd.DataFrame({
            'Coefficient': model.params,
            'Std. Error': model.bse,
            't-value': model.tvalues,
            'p-value': model.pvalues
        })
        return self.results_df

    def find_beta(self):
        beta_messages = []  # Initialize an empty list to store beta messages
        beta_data = []
        for years in self.periods_in_years:
            if years < self.years_difference:
                periods = years * 12
                X = self.monthly_returns[f'{self.benchmark} - RF'][-periods:]
                y = self.monthly_returns[f'{self.ticker} - RF'][-periods:]
                
                slope = np.polyfit(X, y, 1)[0]
                
                beta_message = f"{years} Year Beta: {slope:.2f}\n"
                beta_messages.append(beta_message)  # Append each message to the list
                            # Append year and beta value as a tuple to the beta_data list
            
            beta_data.append((f"{years} Year Beta", slope))

        # Convert list of tuples into a pandas DataFrame
        self.beta_df = pd.DataFrame(beta_data, columns=['Year', 'Beta'])
        # Combine all messages into a single string, separated by newlines
        combined_messages = "\n".join(beta_messages)
        return combined_messages

    def save_to_excel(self, file_path=None):
        if file_path is None:
            directory = os.path.join('.', self.ticker) 
            os.makedirs(directory, exist_ok=True)
            file_path = os.path.join(directory, f'{self.ticker}_{self.word}_Data_CAPM_Analysis_{self.today}.xlsx')
        
        image_path = os.path.join(directory, "regression_plot.png")
        self.plot_regression(save_image=True, image_path=image_path)

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            self.results_df.to_excel(writer, sheet_name='OLS Results')
            self.monthly_returns.to_excel(writer, sheet_name='Completed Data')
            self.ticker_df.to_excel(writer, sheet_name=self.ticker)
            self.benchmark_df.to_excel(writer, sheet_name=self.benchmark)
            self.rf_df.to_excel(writer, sheet_name=self.risk_free)
            self.beta_df.to_excel(writer, sheet_name='Beta Analysis')

        workbook = load_workbook(file_path)
        
        # if 'OLS Results' in workbook.sheetnames:
        #     sheet = workbook['OLS Results']
        # else:
        #     sheet = workbook.create_sheet('OLS Results')
        # img = Image(image_path)
        # sheet.add_image(img, 'G2')

        workbook.save(file_path)
        return file_path

    def plot_regression(self, save_image=True, image_path=None):
        X = self.monthly_returns[f'{self.benchmark} - RF']
        y = self.monthly_returns[f'{self.ticker} - RF']
        
        beta = self.results_df.loc[f'{self.benchmark} - RF', 'Coefficient']
        alpha = self.results_df.loc['const', 'Coefficient']
        x_vals = np.linspace(X.min(), X.max(), 100)
        y_vals = beta * x_vals + alpha
        
        fig = go.Figure(data=go.Scatter(x=X, y=y, mode='markers', name='Data Points'))
        fig.add_trace(go.Scatter(x=x_vals, y=y_vals, mode='lines', name='Regression Line'))
        fig.update_layout(title=f'{self.word} Returns: Asset vs. Benchmark',
                          xaxis_title=f'{self.benchmark} - RF',
                          yaxis_title=f'{self.ticker} - RF')
        
        if save_image:
            if image_path is None:
                image_path = f"{self.ticker}/regression_plot.png"
            os.makedirs(os.path.dirname(image_path), exist_ok=True)
            fig.write_image(image_path)

        return fig, image_path
    


    def upload_to_web(self, image_path):
        with open(image_path, 'rb') as image:
            response = requests.post(
                'https://api.imgbb.com/1/upload',
                data={'key': '403c33d84032226e5659348864f8a16e'},
                files={'image': image}
            )
            if response.status_code == 200:
                return response.json()['data']['url']
            else:
                return None



    
    def calculate_rolling_beta(self):
        """
        Calculate rolling betas for 12, 6, and 3 month windows.
        """
        self.window_sizes = [60, 24, 12, 6] #,3]
        

        X = sm.add_constant(self.monthly_returns[f'{self.benchmark} - RF'])
        y = self.monthly_returns[f'{self.ticker} - RF']
        
        for window_size in self.window_sizes:
            rolling_beta = pd.Series(index=self.monthly_returns.index, dtype=float)
            
            for end in range(window_size, len(self.monthly_returns) + 1):
                X_window = X.iloc[end-window_size:end]
                y_window = y.iloc[end-window_size:end]
                
                model = sm.OLS(y_window, X_window).fit()
                
                # Store the beta coefficient (slope) in the rolling_beta series
                rolling_beta.iloc[end-1] = model.params[1]  # Index 1 because 0 is the intercept
            
            # Store each rolling beta calculation in the DataFrame
            self.monthly_returns[f'Rolling Beta {window_size}M'] = rolling_beta
        
  



    def plot_rolling_beta(self, save_image=True, image_path=None):
        """
        Plot the rolling betas for 12, 6, and 3 month windows using Plotly.
        """
        fig = go.Figure()
        
        for window_size in self.window_sizes:
            column_name = f'Rolling Beta {window_size}M'
            if column_name in self.monthly_returns.columns:
                fig.add_trace(go.Scatter(x=self.monthly_returns.index, y=self.monthly_returns[column_name],
                                        mode='lines', name=f'Rolling Beta {window_size}M'))
            else:
                print(f"{column_name} not calculated. Please run calculate_rolling_beta first.")
        
        fig.update_layout(title=f'Rolling Beta Comparison of {self.ticker} relative to {self.benchmark}',
                        xaxis_title='Date',
                        yaxis_title='Rolling Beta',
                        template='plotly_dark')  # Adjust the template as needed

        if save_image:
            if image_path is None:
                image_path = f"{self.ticker}/rollingbeta_plot.png"
            os.makedirs(os.path.dirname(image_path), exist_ok=True)
            fig.write_image(image_path)
        return fig, image_path
    
        

if __name__ == "__main__":
    benchmark = input('Enter your benchmark ticker (^GSPC): ')
    ticker = input('Enter your stock ticker: ')
    risk_free = input('Enter your risk free ticker (^TYX): ')
    daily_or_monthly = input('Do you want daily or monthly data? (Enter D or M): ')

    analysis = CAPMAnalysis(ticker, benchmark, risk_free, daily_or_monthly)
    analysis.download_data()
    analysis.calculate_returns()
    analysis.perform_regression()
    analysis.find_beta()
    fig1, image_path1 = analysis.plot_regression()
    analysis.calculate_rolling_beta()  # For example, a 12-month rolling window
    fig2, image_path2 = analysis.plot_rolling_beta()
    analysis.save_to_excel()
    fig1.show()
    fig2.show()